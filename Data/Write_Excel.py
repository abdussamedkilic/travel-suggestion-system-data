from openpyxl import Workbook
from pkg_resources import working_set
import xlsxwriter

class WriteExcel:

    workbook_doc2vec = xlsxwriter.Workbook("Output/Doc2vec_output.xlsx")
    workbook_cnn = xlsxwriter.Workbook("Output/CNN_output.xlsx")
    workbook_bert = xlsxwriter.Workbook("Output/Bert_output.xlsx")
    workbook_merged = xlsxwriter.Workbook("Output/output.xlsx")#merged outputs

    def __init__(self) -> None:
        pass
    
    def writeExcel_Doc2vec(self,similarity_matrix,placeName_list,city_name):
        # similarity_matrix's size = (place number , place number)

        #workbook = xlsxwriter.Workbook("Output/Doc2vec_output.xlsx")
        worksheet_output = self.workbook_doc2vec.add_worksheet(city_name)
        
        for i in range(0,len(placeName_list[0])): #place number
            worksheet_output.write(i+1,0,placeName_list[0][i])
            worksheet_output.write(0,i+1,placeName_list[0][i])
            for j in range(0,len(placeName_list[0])):
                worksheet_output.write(i+1,j+1,similarity_matrix[i][j])

        print(" writing is completed for Doc2Vec")
        #self.workbook_doc2vec.close()

    def writeExcel_CNN(self,image_list,city_name,similarity_matrix):
        # preprocess images name
        for i in range(0,len(image_list)):
            image_list[i] = image_list[i].replace('.jpg','')
        
        # write excel
        #workbook = xlsxwriter.Workbook("Output/CNN_output.xlsx")
        worksheet_output = self.workbook_cnn.add_worksheet(city_name)
        
        for i in range(0,len(image_list)): #place number
            worksheet_output.write(i+1,0,image_list[i])
            worksheet_output.write(0,i+1,image_list[i])
            for j in range(0,len(image_list)): #placenumber
                worksheet_output.write(i+1,j+1,similarity_matrix[i][j])
        
        print(" writing is completed for CNN")
        #self.workbook_cnn.close()

    def writeExcel_Bert(self,similarity_matrix,placeName_list,city_name):
        # similarity_matrix's size = (place number , place number)

        #workbook = xlsxwriter.Workbook("Output/Bert_output.xlsx")
        worksheet_output = self.workbook_bert.add_worksheet(city_name)
        
        for i in range(0,len(placeName_list[0])): #place number
            worksheet_output.write(i+1,0,placeName_list[0][i])
            worksheet_output.write(0,i+1,placeName_list[0][i])
            for j in range(0,len(placeName_list[0])):
                worksheet_output.write(i+1,j+1,similarity_matrix[i][j])

        print(" writing is completed for Bert")
        #self.workbook_bert.close()

    def writeExcel_MergedResult(self,similarity_matrix,placeName_list,city_name):
        
        #workbook = xlsxwriter.Workbook("Output/output.xlsx")
        
        worksheet_output = self.workbook_merged.add_worksheet(city_name)

        for i in range(0,len(placeName_list[0])): #place number
            worksheet_output.write(i+1,0,placeName_list[0][i])
            worksheet_output.write(0,i+1,placeName_list[0][i])
            for j in range(0,len(placeName_list[0])):
                worksheet_output.write(i+1,j+1,similarity_matrix[i][j])

        print(" writing is completed for Merged Matrix")
        #self.workbook_merged.close()