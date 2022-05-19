import os
import openpyxl
from sqlalchemy import null

class UpdateExcel:
    
    files = os.listdir("Output/")

    loaded_workbook_doc2vec = null
    loaded_workbook_cnn = null
    loaded_workbook_bert = null
    loaded_workbook_merged = null
    
    def __init__(self) -> None:
        pass

    def control_file(self,file_name):
        
        for file in self.files:
             if file == file_name:
                self.loaded_workbook_doc2vec = openpyxl.load_workbook("Output/Doc2vec_output.xlsx")
                self.loaded_workbook_cnn = openpyxl.load_workbook("Output/CNN_output.xlsx")
                self.loaded_workbook_bert = openpyxl.load_workbook("Output/Bert_output.xlsx")
                self.loaded_workbook_merged = openpyxl.load_workbook("Output/output.xlsx")  # merged outputs
                return True
        return False

    def update_excel_doc2vec(self, similarity_matrix, placeName_list, city_name): 
        
        sheet = self.loaded_workbook_doc2vec.create_sheet(city_name)

        for i in range(0, len(placeName_list[0])):  # place number
            sheet.cell(row=i+2,column=1).value=placeName_list[0][i]
            sheet.cell(row=1,column=i+2).value=placeName_list[0][i]
            for j in range(0, len(placeName_list[0])):  # place number
                sheet.cell(row=i + 2, column=j + 2).value=similarity_matrix[i][j]

        print("update is completed for Doc2vec")

    def update_excel_cnn(self, image_list, city_name, similarity_matrix):
        
        # preprocess images name
        for i in range(0, len(image_list)):
            image_list[i] = image_list[i].replace(".jpg", "")

        sheet = self.loaded_workbook_cnn.create_sheet(city_name)

        for i in range(0, len(image_list)):  # place number
            sheet.cell(row=i+2,column=1).value=image_list[i]
            sheet.cell(row=1,column=i+2).value=image_list[i]
            for j in range(0, len(image_list)):  # place number
                sheet.cell(row=i + 2, column=j + 2).value=similarity_matrix[i][j]

        print("update is completed for CNN")

    def update_excel_bert(self, similarity_matrix, placeName_list, city_name):
        
        sheet = self.loaded_workbook_bert.create_sheet(city_name)

        for i in range(0, len(placeName_list[0])):  # place number
            sheet.cell(row=i+2,column=1).value=placeName_list[0][i]
            sheet.cell(row=1,column=i+2).value=placeName_list[0][i]
            for j in range(0, len(placeName_list[0])):  # place number
                sheet.cell(row=i + 2, column=j + 2).value=similarity_matrix[i][j]
        
        print("update is completed for Bert")

    def update_excel_merged(self, similarity_matrix, placeName_list, city_name):
        
        sheet = self.loaded_workbook_merged.create_sheet(city_name)

        for i in range(0, len(placeName_list)):  # place number
            sheet.cell(row=i+2,column=1).value=placeName_list[i]
            sheet.cell(row=1,column=i+2).value=placeName_list[i]
            for j in range(0, len(placeName_list)):  # place number
                sheet.cell(row=i + 2, column=j + 2).value=similarity_matrix[i][j]

        print("update is completed for Merged Result")

        

